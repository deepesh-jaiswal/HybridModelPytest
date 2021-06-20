import openpyxl

def getRowCount(filepath,Sheet):
    workbook=openpyxl.load_workbook(filepath)
    sheet=workbook.get_sheet_by_name(Sheet)
    return sheet.max_row

def getColumnCount(filepath,Sheet):
    workbook=openpyxl.load_workbook(filepath)
    sheet=workbook.get_sheet_by_name(Sheet)
    return sheet.max_column

def readDataFile(filepath,Sheet,rowno,colno):
    workbook=openpyxl.load_workbook(filepath)
    sheet=workbook.get_sheet_by_name(Sheet)
    return sheet.cell(row=rowno,column=colno).value

def writeDataFile(filepath,Sheet,rowno,colno,data):
    workbook=openpyxl.load_workbook(filepath)
    sheet=workbook.get_sheet_by_name(Sheet)
    sheet.cell(row=rowno,column=colno).value=data
    workbook.save(filepath)      

def createDataDictionary(filepath,Sheet):
    resultant = {}
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.get_sheet_by_name(Sheet)
    for columns in range(1,getColumnCount(filepath,Sheet)+1):
        cell_key = sheet.cell(row=1,column=columns).value
        li = []
        li.append(sheet.cell(row=2,column=columns).value)
        if sheet.cell(row=1,column=1).value.upper() == "iteration".upper():
            for rows in range(3,getRowCount(filepath,Sheet)+1):
                if sheet.cell(row=rows,column=1).value == rows-1:
                    li.append(sheet.cell(row=rows,column=columns).value)
        resultant[cell_key] = li
    return resultant